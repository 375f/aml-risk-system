"""
tests/test_parser.py — тесты модуля module1.parser.

Покрываемые сценарии:
  - Формат А: отдельные колонки Дебет / Кредит
  - Формат Б: Сумма + Тип операции
  - XLSX-файл с авто-определением листа и строки-заголовка
  - Fuzzy-matching нестандартных заголовков
  - Кодировка CP1251
  - Несколько форматов дат
  - Нормализация сумм (пробелы, запятые, символы валюты)
  - Нормализация ИНН (КПП, float, нестандартная длина)
  - Вычисление периода (date_from / date_to)
  - Ошибки: пустой файл, отсутствие даты, отсутствие сумм
"""
import io

import openpyxl
import pandas as pd
import pytest

from module1.parser import (
    ColumnMappingError,
    ParseError,
    _map_columns,
    _score_header,
    parse_statement,
)
from tests.conftest import csv_buf


# ---------------------------------------------------------------------------
# Хелперы
# ---------------------------------------------------------------------------

def xlsx_buf(headers: list, rows: list) -> io.BytesIO:
    """Создать in-memory XLSX."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "test.xlsx"
    return buf


# Стандартное содержимое CSV Формат А (Дебет / Кредит)
_CSV_A = "\n".join([
    "Дата;Назначение платежа;Дебет;Кредит;Контрагент;ИНН контрагента",
    "01.03.2024;Снятие наличных;50000;;Банкомат;",
    "05.03.2024;Оплата поставщику;;100000;ООО Ромашка;7701234560",
    "10.03.2024;Налог НДС;5000;;ИФНС;7707289922",
    "15.03.2024;Зарплата сотруднику;30000;;Петров И.И.;",
])

# Стандартное содержимое CSV Формат Б (Сумма + Тип)
_CSV_B = "\n".join([
    "Дата операции;Сумма;Тип операции;Описание;Контрагент",
    "01.03.2024;50000;Расход;Снятие наличных;Банкомат",
    "05.03.2024;100000;Приход;Оплата услуг;ООО Гамма",
    "10.03.2024;5000;Расход;Налог НДС;ИФНС",
    "15.03.2024;30000;Расход;Зарплата;Иванов И.И.",
])


# ---------------------------------------------------------------------------
# Формат А: Дебет / Кредит
# ---------------------------------------------------------------------------

class TestFormatA:

    def test_returns_dataframe(self):
        df, _, _ = parse_statement(csv_buf(_CSV_A))
        assert isinstance(df, pd.DataFrame)

    def test_canonical_columns_present(self):
        df, _, _ = parse_statement(csv_buf(_CSV_A))
        expected = {"date", "amount", "type", "description", "counterparty", "inn"}
        assert expected.issubset(set(df.columns))

    def test_row_count(self):
        # 3 дебета + 1 кредит = 4 строки
        df, _, _ = parse_statement(csv_buf(_CSV_A))
        assert len(df) == 4

    def test_type_values_only_debit_credit(self):
        df, _, _ = parse_statement(csv_buf(_CSV_A))
        assert set(df["type"].unique()).issubset({"debit", "credit"})

    def test_debit_credit_split_correct(self):
        df, _, _ = parse_statement(csv_buf(_CSV_A))
        assert df[df["type"] == "debit"].shape[0] == 3
        assert df[df["type"] == "credit"].shape[0] == 1

    def test_amounts_are_positive(self):
        df, _, _ = parse_statement(csv_buf(_CSV_A))
        assert (df["amount"] > 0).all()

    def test_period_date_from(self):
        _, date_from, _ = parse_statement(csv_buf(_CSV_A))
        assert date_from is not None
        assert date_from.day == 1 and date_from.month == 3

    def test_period_date_to(self):
        _, _, date_to = parse_statement(csv_buf(_CSV_A))
        assert date_to is not None
        assert date_to.day == 15 and date_to.month == 3

    def test_inn_preserved(self):
        df, _, _ = parse_statement(csv_buf(_CSV_A))
        assert "7701234560" in df["inn"].values
        assert "7707289922" in df["inn"].values

    def test_sorted_by_date(self):
        df, _, _ = parse_statement(csv_buf(_CSV_A))
        dates = pd.to_datetime(df["date"])
        assert (dates.diff().dropna() >= pd.Timedelta(0)).all()


# ---------------------------------------------------------------------------
# Формат Б: Сумма + Тип операции
# ---------------------------------------------------------------------------

class TestFormatB:

    def test_returns_dataframe(self):
        df, _, _ = parse_statement(csv_buf(_CSV_B))
        assert isinstance(df, pd.DataFrame)

    def test_row_count(self):
        df, _, _ = parse_statement(csv_buf(_CSV_B))
        assert len(df) == 4

    def test_type_classification_from_text(self):
        df, _, _ = parse_statement(csv_buf(_CSV_B))
        # "Приход" → credit; "Расход" → debit
        assert df[df["type"] == "credit"].shape[0] == 1
        assert df[df["type"] == "debit"].shape[0] == 3

    def test_amounts_from_single_column(self):
        df, _, _ = parse_statement(csv_buf(_CSV_B))
        assert set(df["amount"].tolist()) == {50_000.0, 100_000.0, 5_000.0, 30_000.0}


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

class TestXlsx:

    def test_basic_xlsx_parse(self):
        headers = ["Дата", "Назначение", "Дебет", "Кредит", "Контрагент"]
        rows = [
            ["01.03.2024", "Снятие наличных", 50_000, None, "Банкомат"],
            ["05.03.2024", "Поступление",      None, 100_000, "ООО А"],
        ]
        df, _, _ = parse_statement(xlsx_buf(headers, rows))
        assert len(df) == 2
        assert set(df["type"].unique()).issubset({"debit", "credit"})

    def test_xlsx_with_metadata_rows(self):
        """XLSX с «шапкой» перед заголовком — parser должен найти строку-заголовок."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Банк: ВТБ"
        ws["A2"] = "Период: 01.03.2024 — 31.03.2024"
        ws.append(["Дата", "Назначение платежа", "Дебет", "Кредит", "Контрагент"])
        ws.append(["01.03.2024", "Снятие наличных", 50_000, None, "Банкомат"])
        ws.append(["05.03.2024", "Поступление",      None, 100_000, "ООО А"])
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0); buf.name = "meta.xlsx"
        df, _, _ = parse_statement(buf)
        assert len(df) == 2


# ---------------------------------------------------------------------------
# Fuzzy-matching заголовков
# ---------------------------------------------------------------------------

class TestFuzzyMapping:

    def test_score_exact_match(self):
        from module1.parser import _ALIASES
        score = _score_header("дата операции", _ALIASES["date"])
        assert score >= 0.95

    def test_score_substring_match(self):
        from module1.parser import _ALIASES
        score = _score_header("инн/кпп контрагента", _ALIASES["inn"])
        assert score >= 0.80

    def test_score_fuzzy_match_debit(self):
        from module1.parser import _ALIASES
        score = _score_header("дебетовый оборот", _ALIASES["debit"])
        assert score >= 0.80

    def test_map_columns_format_a(self):
        mapping = _map_columns(["Дата", "Назначение платежа", "Дебет", "Кредит", "Контрагент"])
        canonical = set(mapping.values())
        assert "date"        in canonical
        assert "description" in canonical
        assert "debit"       in canonical
        assert "credit"      in canonical

    def test_map_columns_format_b(self):
        mapping = _map_columns(["Дата операции", "Сумма", "Тип операции", "Описание"])
        canonical = set(mapping.values())
        assert "date"        in canonical
        assert "amount"      in canonical
        assert "type"        in canonical
        assert "description" in canonical

    def test_map_columns_tinkoff_style(self):
        """Заголовки, характерные для выписки Тинькофф."""
        mapping = _map_columns([
            "Дата проводки", "Сумма операции", "Тип", "Описание операции", "Контрагент"
        ])
        canonical = set(mapping.values())
        assert "date"   in canonical
        assert "amount" in canonical

    def test_no_duplicate_canonical_assignments(self):
        """Каждое каноническое имя назначается не более одному сырому заголовку."""
        cols = ["Дата", "Дата операции", "Сумма", "Дебет", "Кредит", "Назначение"]
        mapping = _map_columns(cols)
        canonical_values = list(mapping.values())
        assert len(canonical_values) == len(set(canonical_values))


# ---------------------------------------------------------------------------
# Кодировка CP1251
# ---------------------------------------------------------------------------

class TestEncoding:

    def test_cp1251_csv(self):
        content = "\n".join([
            "Дата;Назначение;Дебет;Кредит;Контрагент",
            "01.03.2024;Снятие наличных;50000;;Банкомат",
            "05.03.2024;Поступление;;100000;ООО Альфа",
        ])
        buf = csv_buf(content, encoding="cp1251")
        df, _, _ = parse_statement(buf)
        assert len(df) == 2
        assert "debit"  in df["type"].values
        assert "credit" in df["type"].values


# ---------------------------------------------------------------------------
# Нормализация дат
# ---------------------------------------------------------------------------

class TestDateParsing:

    def _parse_single(self, date_str: str) -> pd.Timestamp:
        content = "\n".join([
            "Дата;Назначение;Сумма;Тип",
            f"{date_str};Оплата;10000;Расход",
        ])
        df, _, _ = parse_statement(csv_buf(content))
        return pd.to_datetime(df.iloc[0]["date"])

    def test_dot_format_dd_mm_yyyy(self):
        ts = self._parse_single("15.06.2024")
        assert ts.day == 15 and ts.month == 6 and ts.year == 2024

    def test_slash_format_dd_mm_yyyy(self):
        ts = self._parse_single("15/06/2024")
        assert ts.day == 15 and ts.month == 6

    def test_iso_format_yyyy_mm_dd(self):
        ts = self._parse_single("2024-06-15")
        assert ts.day == 15 and ts.month == 6

    def test_datetime_with_time(self):
        ts = self._parse_single("15.06.2024 14:30")
        assert ts.day == 15 and ts.month == 6


# ---------------------------------------------------------------------------
# Нормализация сумм
# ---------------------------------------------------------------------------

class TestAmountNormalization:

    def _parse_amount(self, amount_str: str) -> float:
        content = "\n".join([
            "Дата;Назначение;Сумма;Тип",
            f"01.03.2024;Тест;{amount_str};Расход",
        ])
        df, _, _ = parse_statement(csv_buf(content))
        return float(df.iloc[0]["amount"])

    def test_integer_amount(self):
        assert self._parse_amount("50000") == pytest.approx(50_000)

    def test_amount_with_spaces(self):
        # Российский формат: «1 000 000,00»
        assert self._parse_amount("1 000 000") == pytest.approx(1_000_000)

    def test_amount_with_decimal_comma(self):
        assert self._parse_amount("50000,50") == pytest.approx(50_000.50)

    def test_negative_amount_becomes_positive(self):
        # Некоторые банки выгружают расходы со знаком минус
        assert self._parse_amount("-30000") == pytest.approx(30_000)


# ---------------------------------------------------------------------------
# Нормализация ИНН
# ---------------------------------------------------------------------------

class TestInnNormalization:

    def _parse_inn(self, inn_val: str) -> str:
        content = "\n".join([
            "Дата;Назначение;Сумма;Тип;ИНН контрагента",
            f"01.03.2024;Оплата;10000;Расход;{inn_val}",
        ])
        df, _, _ = parse_statement(csv_buf(content))
        return str(df.iloc[0]["inn"])

    def test_valid_10_digit_inn(self):
        assert self._parse_inn("7701234560") == "7701234560"

    def test_valid_12_digit_inn(self):
        assert self._parse_inn("770123456789") == "770123456789"

    def test_inn_with_kpp_stripped(self):
        # "7701234560/770101001" → только ИНН
        assert self._parse_inn("7701234560/770101001") == "7701234560"

    def test_inn_as_float_from_excel(self):
        # Excel читает ИНН как число: 7701234560.0
        assert self._parse_inn("7701234560.0") == "7701234560"

    def test_invalid_inn_length_returns_empty(self):
        # 9 цифр — невалидный ИНН
        assert self._parse_inn("123456789") == ""


# ---------------------------------------------------------------------------
# Обработка ошибок
# ---------------------------------------------------------------------------

class TestErrors:

    def test_empty_file_raises_parse_error(self):
        buf = csv_buf("Дата;Сумма\n")  # 0 строк данных
        with pytest.raises(ParseError):
            parse_statement(buf)

    def test_no_date_column_raises_column_mapping_error(self):
        buf = csv_buf("Товар;Цена;Количество\nЯблоко;100;5\nГруша;80;3\n")
        with pytest.raises((ParseError, ColumnMappingError)):
            parse_statement(buf)

    def test_no_amount_column_raises_column_mapping_error(self):
        buf = csv_buf("Дата;Описание;Контрагент\n01.03.2024;Платёж;ООО А\n")
        with pytest.raises((ParseError, ColumnMappingError)):
            parse_statement(buf)

    def test_all_zero_amounts_raises_parse_error(self):
        content = "\n".join([
            "Дата;Назначение;Сумма;Тип",
            "01.03.2024;Тест;0;Расход",
            "02.03.2024;Тест2;0;Приход",
        ])
        with pytest.raises(ParseError):
            parse_statement(csv_buf(content))

    def test_unsupported_extension_raises_parse_error(self):
        buf = io.BytesIO(b"garbage")
        buf.name = "data.pdf"
        with pytest.raises(ParseError):
            parse_statement(buf)

    def test_only_credits_no_debits_is_valid(self):
        """Выписка только с поступлениями — не ошибка, должна разобраться."""
        content = "\n".join([
            "Дата;Назначение;Сумма;Тип",
            "01.03.2024;Поступление;100000;Приход",
            "05.03.2024;Поступление;50000;Приход",
        ])
        df, _, _ = parse_statement(csv_buf(content))
        assert len(df) == 2
        assert (df["type"] == "credit").all()
